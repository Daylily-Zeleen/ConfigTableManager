import subprocess
import logging


# 检查
_result = subprocess.run("pip show openpyxl", shell=True, capture_output=True)
if _result.returncode != 0:
    logging.error("This tool is require 'openpyxl' package, please run: pip install openpyxl")
    exit(1)


import sys
import json
import os
import openpyxl
from openpyxl import Workbook


def dump_xlsx_to_json(xlsx_file_path: str, output_json_path: str):
    workbook: Workbook = openpyxl.load_workbook(xlsx_file_path)

    data: dict[str, dict] = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Sheet
        sheet_data: list[list[dict[str, None]]] = []
        for row in range(sheet.max_row):
            row_data: list[dict] = []
            for column in range(sheet.max_column):
                cell = sheet.cell(row=row + 1, column=column + 1)
                row_data.append(
                    {
                        "value": cell.value,
                        # TODO: 存储更多的单元格属性
                    }
                )
            sheet_data.append(row_data)

        data[sheet_name] = {
            "data": sheet_data,
            "sheet_format": _to_dict(sheet.sheet_format),
            # TODO: 存储更多的Sheet属性
        }

    workbook.close()
    f = open(output_json_path, "w", encoding="utf8")
    json.dump(obj=data, fp=f, indent="\t")
    f.close()
    return True


def override_xlsx_worksheets_from_json(json_file_path: str, output_xlsx_file_path: str):
    f = open(json_file_path, "r", encoding="utf8")
    data: dict[str, dict] = json.load(f)

    if len(data) <= 0:
        logging.error("The input json has not data.")
        exit(1)

    workbook: Workbook
    default_sheetnames: list[str] = []
    if os.path.exists(output_xlsx_file_path):
        workbook = openpyxl.load_workbook(output_xlsx_file_path)
    else:
        workbook = Workbook()
        default_sheetnames = workbook.sheetnames

    for sheet_name in data:
        if sheet_name in default_sheetnames:
            default_sheetnames.remove(sheet_name)

        # 覆盖
        if not sheet_name in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
        sheet = workbook[sheet_name]

        if "sheet_format" in data[sheet_name]:
            sheet_format: dict = data[sheet_name]["sheet_format"]
            for k in sheet_format:
                setattr(sheet.sheet_format, k, sheet_format[k])
        # TODO 其他表格属性

        # 数据行
        sheet_data: list[list[dict[str, None]]] = data[sheet_name]["data"]
        for row in range(len(sheet_data)):
            for column in range(len(sheet_data[row])):
                cell_data: dict[str, None] = sheet_data[row][column]
                sheet.cell(row=row + 1, column=column + 1, value=cell_data["value"])
                # TODO: 其他单元格属性(先检查是否有对应属性)
            # 清除右侧的多余格
            for column in range(len(sheet_data[row]), sheet.max_column):
                cell = sheet.cell(row=row + 1, column=column + 1)
                if not cell.value is None:
                    cell.value = ""

    for sheet_name in default_sheetnames:
        if sheet_name in workbook:
            del workbook[sheet_name]

    workbook.save(output_xlsx_file_path)
    workbook.close()


def _to_dict(obj: None) -> dict:
    return {key: getattr(obj, key) for key in vars(obj) if not key.startswith("_")}


def main(argv):
    for arg in argv:
        if arg in ["-h", "--help"]:
            print("Valid command:")
            print("\tpython xlsx_json.py --dump_json path/to/excel.xlsx path/to/output.json")
            print("\tpython xlsx_json.py --override_xlsx path/to/data.json path/to/output.xlsx")
            exit(0)

    invalid_hint = "Invalid arguments for 'xlsx_json' tool. Please type '-h' or '--help' for more details."
    if len(argv) != 3:

        logging.error(invalid_hint)
        exit(1)

    if argv[0] == "--dump_json":
        dump_xlsx_to_json(argv[1], argv[2])
    elif argv[0] == "--override_xlsx":
        override_xlsx_worksheets_from_json(argv[1], argv[2])
    else:

        logging.error(invalid_hint)
        exit(1)


if __name__ == "__main__":
    main(sys.argv[1:])
